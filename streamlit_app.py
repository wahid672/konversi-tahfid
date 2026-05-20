import streamlit as st
import pandas as pd
import datetime
import re
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fungsi Konversi Waktu
def hijri_to_gregorian(y, m, d):
    jd = int((11 * y + 3) / 30) + 354 * y + 30 * m - int((m - 1) / 2) + d + 1948440 - 385
    if jd > 2299160:
        l = jd + 68569
        n = int((4 * l) / 146097)
        l = l - int((146097 * n + 3) / 4)
        i = int((4000 * (l + 1)) / 1461001)
        l = l - int((1461 * i) / 4) + 31
        j = int((80 * l) / 2447)
        day = l - int((2447 * j) / 80)
        l = int(j / 11)
        month = j + 2 - 12 * l
        year = 100 * (n - 49) + i + l
    else:
        j = jd + 1402
        k = int((j - 1) / 1461)
        l = j - 1461 * k
        n = int((l - 1) / 365) - int(l / 1461)
        i = l - 365 * n + 30
        j = int((80 * i) / 2447)
        day = i - int((2447 * j) / 80)
        i = int(j / 11)
        month = j + 2 - 12 * i
        year = 4 * k + n + i - 4716
    return datetime.date(year, month, day)

def convert_waktu(waktu_str):
    try:
        parts = str(waktu_str).split(',')
        if len(parts) >= 2:
            date_time = parts[1].strip()
            dt_parts = date_time.split(' ')
            date_str = dt_parts[0]
            time_str = dt_parts[1] if len(dt_parts) > 1 else ""
            d, m, y = map(int, date_str.split('/'))
            greg_date = hijri_to_gregorian(y, m, d)
            return f"{greg_date.strftime('%d/%m/%Y')} {time_str}".strip()
    except Exception:
        pass
    return waktu_str

def extract_nilai(rincian_str):
    try:
        match = re.search(r'Ket\s*:\s*(\d+)', str(rincian_str), re.IGNORECASE)
        if match:
            return int(match.group(1))
    except:
        pass
    return None

# --- TAMPILAN WEB STREAMLIT ---
st.set_page_config(page_title="Konversi Laporan Tahfidz PPTQ DAMUS", layout="centered")

st.title("Aplikasi Konversi Laporan Tahfidz")
st.write("Upload file **Excel (.xlsx)** laporan tahfidz dari siakad pptqdamus.com, dan unduh versi Excel yang sudah dirapikan.")

# UBAH TIPE FILE MENJADI XLSX
uploaded_file = st.file_uploader("Pilih file Excel Laporan", type=["xlsx"])

if uploaded_file is not None:
    st.info("Memproses data...")
    try:
        # MENGGUNAKAN pd.read_excel BUKAN pd.read_csv
        df = pd.read_excel(uploaded_file, skiprows=6)
        
        df['WAKTU_MASEHI'] = df['WAKTU'].apply(convert_waktu)
        df['NILAI'] = df['RINCIAN'].apply(extract_nilai)
        
        cols = list(df.columns)
        cols.insert(cols.index('WAKTU') + 1, cols.pop(cols.index('WAKTU_MASEHI')))
        cols.insert(cols.index('RINCIAN') + 1, cols.pop(cols.index('NILAI')))
        df = df[cols]
        df.columns = ['NO', 'WAKTU (HIJRI)', 'WAKTU (MASEHI)', 'NAMA SANTRI', 'RINCIAN LAPORAN', 'NILAI', 'USTADZ/USTADZAH']

        # Proses OpenPyXL (Styling)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laporan Tahfidz"

        ws['A1'] = "LAPORAN TAHFIDZ"
        ws['A1'].font = Font(name="Arial", size=16, bold=True, color="1F497D")
        ws['A2'] = "PON. PES. TAHFIDZ ANAK DARUL MUSTHOFA"
        ws['A2'].font = Font(name="Arial", size=12, bold=True, color="595959")
        ws['A3'] = "Kp. Krajan RT 02 RW 04 Jogoloyo, Wonosalam, Demak, Jawa Tengah Telp. 081390540000"
        ws['A3'].font = Font(name="Arial", size=10, italic=True, color="7F7F7F")
        ws['A4'] = "Periode: Laporan Hasil Konversi"
        ws['A4'].font = Font(name="Arial", size=10, color="595959")

        header_row = 7
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
        zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for row_idx, row_data in enumerate(df.values, start=header_row + 1):
            is_even = (row_idx % 2 == 0)
            current_fill = zebra_fill if is_even else white_fill
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Arial", size=10)
                cell.border = thin_border
                cell.fill = current_fill
                
                col_name = headers[col_idx - 1]
                if col_name in ['NO', 'WAKTU (HIJRI)', 'WAKTU (MASEHI)', 'NILAI']:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name == 'RINCIAN LAPORAN':
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            col_name = ws.cell(row=header_row, column=col[0].column).value
            if col_name == 'NO': ws.column_dimensions[col_letter].width = 6
            elif col_name == 'NILAI': ws.column_dimensions[col_letter].width = 10
            elif col_name == 'RINCIAN LAPORAN': ws.column_dimensions[col_letter].width = 45
            elif col_name and 'WAKTU' in col_name: ws.column_dimensions[col_letter].width = 25
            else: ws.column_dimensions[col_letter].width = 35

        ws.row_dimensions[header_row].height = 28
        for r in range(header_row + 1, ws.max_row + 1):
            rincian_val = ws.cell(row=r, column=5).value
            if rincian_val and '\n' in str(rincian_val):
                num_lines = str(rincian_val).count('\n') + 1
                ws.row_dimensions[r].height = max(18 * num_lines, 22)
            else:
                ws.row_dimensions[r].height = 22

        # Simpan ke dalam buffer memori agar bisa didownload
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Konversi Excel berhasil!")
        
        # Tombol Download
        st.download_button(
            label="⬇️ Download File Excel Hasil (.xlsx)",
            data=output,
            file_name="Laporan_Tahfidz_Hasil.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Tampilkan sedikit preview data di web
        st.write("Preview Data:")
        st.dataframe(df.head())

    except Exception as e:
        st.error(f"Terjadi kesalahan saat memproses file Excel: {e}")
